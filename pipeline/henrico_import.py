"""
henrico_import.py — Monthly Henrico County permit import.

Downloads the monthly building permit Excel from henrico.gov,
filters for target ZIPs and luxury keywords, inserts new records
as Queued with source='Henrico Direct'.

No Shovels data for Henrico — no assessed value, owner type, or
contact enrichment available from this source.

Run:  python -m pipeline.henrico_import
Cron: 0 8 5 * *  (5th of each month, 8:00 AM ET)
"""
import io
import logging
from datetime import date

import httpx
import openpyxl

import config
import db
from pipeline.monday_pull import build_purl_url

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s — %(message)s",
)
log = logging.getLogger(__name__)

CUSTOMER_ID = "livewire"


def build_excel_url(for_date: date = None) -> str:
    """Build the henrico.gov Excel URL for the given month (defaults to current month)."""
    d = for_date or date.today()
    mon  = d.strftime("%b").upper()   # e.g. "MAR"
    year = d.strftime("%Y")            # e.g. "2026"
    return config.HENRICO_EXCEL_URL.replace("{MON}", mon).replace("{YEAR}", year)


def download_excel(url: str) -> bytes | None:
    try:
        r = httpx.get(url, timeout=60, follow_redirects=True)
        r.raise_for_status()
        log.info("Downloaded Henrico Excel (%d bytes) from %s", len(r.content), url)
        return r.content
    except httpx.HTTPError as e:
        log.error("Failed to download Henrico Excel from %s: %s", url, e)
        return None


def keyword_match(text: str) -> bool:
    if not text:
        return False
    text_lower = text.lower()
    return any(kw in text_lower for kw in config.HENRICO_KEYWORDS)


def normalize_address(row_address: str, city: str = "", state: str = "VA", zip_code: str = "") -> str:
    parts = [p for p in [row_address, city, state, zip_code] if p]
    return ", ".join(parts).strip()


def parse_and_import(content: bytes) -> tuple[int, int]:
    """
    Parse the Excel workbook and insert qualifying records.
    Returns (rows_checked, rows_inserted).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # Find the header row (first row containing 'ZIP' or 'zip' or 'Zip')
    headers = None
    header_row_idx = 0
    for idx, row in enumerate(rows_iter):
        row_strs = [str(c).lower() if c else "" for c in row]
        if any("zip" in c for c in row_strs) or any("permit" in c for c in row_strs):
            headers = [str(c).strip().lower() if c else "" for c in row]
            header_row_idx = idx
            log.info("Header row found at index %d: %s", idx, headers)
            break

    if headers is None:
        # Fallback: assume first row is header
        wb2 = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws2 = wb2.active
        all_rows = list(ws2.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else f"col{i}" for i, c in enumerate(all_rows[0])]
        data_rows = all_rows[1:]
        log.info("Fallback: using first row as header: %s", headers)
    else:
        # Consume remaining rows from the same iterator
        data_rows = list(rows_iter)

    # Map common column name variants
    def col(name_variants: list[str]) -> int | None:
        for v in name_variants:
            for i, h in enumerate(headers):
                if v in h:
                    return i
        return None

    idx_zip       = col(["zip"])
    idx_address   = col(["address", "location", "site"])
    idx_desc      = col(["description", "work", "job", "type", "permit type"])
    idx_date      = col(["date", "issue", "filed", "permit date"])
    idx_owner     = col(["owner", "applicant", "name"])
    idx_contractor= col(["contractor", "builder"])
    idx_value     = col(["value", "valuation", "cost", "fee"])
    idx_permit_no = col(["permit", "number", "no", "#"])

    log.info("Column map — zip:%s address:%s desc:%s date:%s owner:%s",
             idx_zip, idx_address, idx_desc, idx_date, idx_owner)

    rows_checked = 0
    rows_inserted = 0

    for row in data_rows:
        if not any(row):
            continue
        rows_checked += 1

        def cell(idx):
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        zip_code = cell(idx_zip)
        if zip_code not in config.HENRICO_ZIPS:
            continue

        description = cell(idx_desc)
        if not keyword_match(description):
            continue

        address_raw  = cell(idx_address)
        owner_name   = cell(idx_owner)
        file_date    = cell(idx_date)
        contractor   = cell(idx_contractor)
        permit_no    = cell(idx_permit_no)

        full_address = normalize_address(address_raw, zip_code=zip_code)
        if not full_address:
            continue

        permit_data = {
            "customer_id":        CUSTOMER_ID,
            "shovels_permit_id":  permit_no or None,
            "source":             "Henrico Direct",
            "property_address":   full_address,
            "property_city":      "Henrico",
            "property_state":     "VA",
            "property_zip":       zip_code,
            "permit_type":        description,
            "permit_tags":        "[]",
            "is_new_construction": 1 if keyword_match(description) and any(
                kw in (description or "").lower() for kw in ["new home", "new house", "single family new"]
            ) else 0,
            "file_date":          file_date,
            "owner_name":         owner_name,
            "owner_type":         "unknown",
            "contractor_name":    contractor,
            "status":             "Queued",
        }

        inserted, permit_id = db.upsert_permit(permit_data)
        if inserted:
            purl = build_purl_url(permit_id)
            db.set_permit_status(permit_id, "Queued", {"purl_url": purl})
            rows_inserted += 1
            log.debug("Inserted Henrico permit: %s", full_address)

    wb.close()
    log.info("Henrico import: checked %d rows, inserted %d new permits.", rows_checked, rows_inserted)
    return rows_checked, rows_inserted


def run(for_date: date = None):
    log.info("=== Henrico Import started ===")
    db.init_db()

    url = build_excel_url(for_date)
    log.info("Excel URL: %s", url)

    content = download_excel(url)
    if not content:
        log.error("No Excel content — aborting.")
        return

    rows_checked, rows_inserted = parse_and_import(content)

    db.set_app_config_field("last_henrico_run", str(date.today()))
    log.info("=== Henrico Import complete — %d inserted ===", rows_inserted)


if __name__ == "__main__":
    run()
