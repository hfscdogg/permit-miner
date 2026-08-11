"""
henrico.py — Henrico County permit scraper (monthly Excel workbooks).

Downloads the henrico.gov monthly building-permit Excel and returns raw
permit dicts for monday_pull.py, which applies the same owner / project /
value filters as every other source — so Henrico records hit the Monday
preview email (with one-click Exclude links) before anything mails.

Henrico publishes one workbook per month, so each run reads the current
and previous month's files (the lookback window can span a month boundary,
and a month's file may only appear after month-end). Rows are filtered to
a window at least MIN_WINDOW_DAYS wide; per-address dedup in monday_pull
and db.upsert_permit makes re-reading the same file idempotent.

Henrico has no public ArcGIS assessor service, so non-new-construction
records qualify via the workbook's valuation column (job_value_dollars)
against config.MIN_JOB_VALUE_DOLLARS.
"""
import io
import logging
from datetime import date, datetime, timedelta

import httpx
import openpyxl

import config

log = logging.getLogger(__name__)

# Henrico's workbook is monthly — keep the row window at least this wide so
# rows published after month-end still fall inside a weekly run's window.
MIN_WINDOW_DAYS = 40


def build_excel_url(for_date: date = None) -> str:
    """Build the henrico.gov Excel URL for the given month (defaults to current month)."""
    d = for_date or date.today()
    mon  = d.strftime("%b").upper()   # e.g. "MAR"
    year = d.strftime("%Y")            # e.g. "2026"
    return config.HENRICO_EXCEL_URL.replace("{MON}", mon).replace("{YEAR}", year)


def download_excel(url: str) -> bytes | None:
    try:
        r = httpx.get(url, timeout=60, follow_redirects=True)
        r.raise_for_status()
        log.info("Downloaded Henrico Excel (%d bytes) from %s", len(r.content), url)
        return r.content
    except httpx.HTTPError as e:
        # Early in the month the current file may not be published yet.
        log.warning("Could not download Henrico Excel from %s: %s", url, e)
        return None


def _parse_file_date(value) -> date | None:
    """Parse an Excel date cell (datetime object or string). None if unparseable."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = str(value).strip().split(" ")[0]
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_dollars(value) -> int:
    """Parse a valuation cell ('$123,456.00', 123456.0, …) into whole dollars."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).replace("$", "").replace(",", "").strip()
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_workbook(content: bytes, cutoff: date) -> list[dict]:
    """Parse one monthly workbook into raw permit dicts (rows dated >= cutoff)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # Find the header row: a multi-column row naming a zip/address column
    # ("permit" alone also appears in title/preamble rows, so it isn't enough).
    headers = None
    for row in rows_iter:
        row_strs = [str(c).lower() if c else "" for c in row]
        populated = sum(1 for c in row_strs if c)
        if populated >= 3 and any("zip" in c or "address" in c for c in row_strs):
            headers = [str(c).strip().lower() if c else "" for c in row]
            log.info("Henrico header row: %s", headers)
            break

    if headers is None:
        # Fallback: assume first row is header
        wb2 = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws2 = wb2.active
        all_rows = list(ws2.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else f"col{i}" for i, c in enumerate(all_rows[0])]
        data_rows = all_rows[1:]
        log.info("Henrico fallback header: %s", headers)
    else:
        data_rows = list(rows_iter)

    # Map common column name variants
    def col(name_variants: list[str]) -> int | None:
        for v in name_variants:
            for i, h in enumerate(headers):
                if v in h:
                    return i
        return None

    idx_zip       = col(["zip"])
    idx_address   = col(["address", "location", "site"])
    idx_desc      = col(["description", "work", "job", "type", "permit type"])
    idx_date      = col(["date", "issue", "filed", "permit date"])
    idx_owner     = col(["owner", "applicant", "name"])
    idx_contractor= col(["contractor", "builder"])
    # Deliberately no "fee" variant — permit fees are not job valuations.
    idx_value     = col(["valuation", "value", "cost"])
    idx_permit_no = col(["permit", "number", "no", "#"])

    records: list[dict] = []
    for row in data_rows:
        if not any(row):
            continue

        def raw(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def cell(idx):
            v = raw(idx)
            return str(v).strip() if v is not None else ""

        zip_code = cell(idx_zip)[:5]
        if zip_code not in config.HENRICO_ZIPS:
            continue

        address = cell(idx_address)
        if not address:
            continue

        # Skip rows older than the window; keep undated rows (dedup protects).
        file_dt = _parse_file_date(raw(idx_date))
        if file_dt and file_dt < cutoff:
            continue

        description = cell(idx_desc)
        records.append({
            "source":             "Henrico Direct",
            "permit_number":      cell(idx_permit_no),
            "permit_type":        description or "Residential Building",
            "property_address":   address,
            "property_city":      "Henrico",
            "property_state":     "VA",
            "property_zip":       zip_code,
            "description":        description,
            "file_date":          file_dt.isoformat() if file_dt else cell(idx_date),
            "job_value_dollars":  _parse_dollars(raw(idx_value)),
            "owner_name":         cell(idx_owner),
            "contractor_name":    cell(idx_contractor),
        })

    wb.close()
    return records


def fetch_permits(since_days: int = 14) -> list[dict]:
    """
    Fetch Henrico permits from the current and previous month's workbooks,
    limited to rows dated within max(since_days, MIN_WINDOW_DAYS) days.
    """
    today = date.today()
    cutoff = today - timedelta(days=max(since_days, MIN_WINDOW_DAYS))
    prev_month = today.replace(day=1) - timedelta(days=1)

    permits: list[dict] = []
    seen: set[str] = set()
    for month in (today, prev_month):
        content = download_excel(build_excel_url(month))
        if not content:
            continue
        for rec in parse_workbook(content, cutoff):
            key = rec["property_address"].lower()
            if key in seen:
                continue
            seen.add(key)
            permits.append(rec)

    log.info("Henrico: %d permits within window (cutoff %s)", len(permits), cutoff)
    return permits


def _debug():
    logging.basicConfig(level=logging.DEBUG)
    for rec in fetch_permits()[:20]:
        print(rec)


if __name__ == "__main__":
    _debug()
